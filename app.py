import sys
import os
import requests
import csv
from datetime import datetime
import threading
from extractor import get_reader

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QFileDialog, QScrollArea,
    QFrame, QProgressBar, QSizePolicy, QDialog, QMessageBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QPoint
from PyQt6.QtGui import QFont, QCursor, QPixmap

from preprocessor import preprocess_image
from extractor import extract_raw_text
from ia import extraire_adresse_ia, ollama_model_exists, ollama_pull_model, ollama_server_reachable, MODEL
from distance import calculer_distance


# ─────────────────────────────────────────────
# Worker thread
# ─────────────────────────────────────────────
class Worker(QThread):
    progress = pyqtSignal(int, str)
    result_ready = pyqtSignal(dict)
    finished = pyqtSignal()

    def __init__(self, image_paths: list[str], adresse_labo: str):
        super().__init__()
        self.image_paths = image_paths
        self.adresse_labo = adresse_labo

    def _log_stage(self, index: int, total: int, filename: str, stage: str) -> None:
        print(f"[{index + 1}/{total}] {filename} - {stage}", flush=True)

    def run(self):
        total = len(self.image_paths)
        for i, img_path in enumerate(self.image_paths):
            filename = os.path.basename(img_path)
            self.progress.emit(i, filename)
            self._log_stage(i, total, filename, "chargement de l'image")

            entry = {
                "fichier": filename,
                "img_path": img_path,
                "adresse": "",
                "distance_km": "",
                "distance_raw": None,
                "confiance": "basse",
                "statut": "erreur",
            }

            try:
                self._log_stage(i, total, filename, "preparation OCR")
                img = preprocess_image(img_path)
                self._log_stage(i, total, filename, "extraction OCR")
                texte = extract_raw_text(img)

                if not texte.strip():
                    entry["statut"] = "vide"
                    self._log_stage(i, total, filename, "aucun texte detecte")
                    self.result_ready.emit(entry)
                    continue

                self._log_stage(i, total, filename, "traitement par l'IA")
                adresse = extraire_adresse_ia(texte)
                if adresse:
                    self._log_stage(i, total, filename, "calcul de la distance")
                    entry["adresse"] = adresse
                    dist, confiance = calculer_distance(self.adresse_labo, adresse)
                    entry["distance_raw"] = dist
                    entry["distance_km"] = f"{dist:.1f} km" if dist else "N/A"
                    entry["confiance"] = confiance
                    entry["statut"] = "ok"
                    self._log_stage(i, total, filename, f"termine -> {entry['distance_km']}")
                else:
                    entry["statut"] = "non_trouve"
                    self._log_stage(i, total, filename, "adresse non trouvee")

            except Exception as e:
                entry["statut"] = "erreur"
                entry["adresse"] = str(e)[:60]
                self._log_stage(i, total, filename, f"erreur -> {entry['adresse']}")

            self.result_ready.emit(entry)

        self.finished.emit()


# ─────────────────────────────────────────────
# Fenetre apercu ticket
# ─────────────────────────────────────────────
class _PannableLabel(QLabel):
    """Image label avec drag-to-pan."""
    def __init__(self, scroll_area: "QScrollArea"):
        super().__init__()
        self._scroll = scroll_area
        self._drag_pos: QPoint | None = None
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))

    def mousePressEvent(self, e):
        if e.button() == Qt.MouseButton.LeftButton:
            self._drag_pos = e.globalPosition().toPoint()
            self.setCursor(QCursor(Qt.CursorShape.ClosedHandCursor))

    def mouseMoveEvent(self, e):
        if self._drag_pos is not None:
            delta = e.globalPosition().toPoint() - self._drag_pos
            self._drag_pos = e.globalPosition().toPoint()
            hbar = self._scroll.horizontalScrollBar()
            vbar = self._scroll.verticalScrollBar()
            hbar.setValue(hbar.value() - delta.x())
            vbar.setValue(vbar.value() - delta.y())

    def mouseReleaseEvent(self, e):
        self._drag_pos = None
        self.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))


class TicketViewer(QDialog):
    def __init__(self, img_path: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Apercu du ticket")
        self.setMinimumSize(500, 700)
        self.setStyleSheet("background: #fff;")
        self._zoom = 1.0
        self._pixmap_orig = QPixmap(img_path)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(8)

        # barre zoom
        zoom_bar = QHBoxLayout()
        zoom_bar.setSpacing(6)

        btn_style = """
            QPushButton {
                background: #f0f0f0; color: #1a1a1a;
                border: 1px solid #ddd; border-radius: 3px;
                padding: 4px 10px; font-family: Georgia; font-size: 12px;
            }
            QPushButton:hover { background: #e0e0e0; }
        """
        btn_minus = QPushButton("−")
        btn_minus.setFixedWidth(32)
        btn_minus.setStyleSheet(btn_style)
        btn_minus.clicked.connect(self._zoom_out)

        btn_plus = QPushButton("+")
        btn_plus.setFixedWidth(32)
        btn_plus.setStyleSheet(btn_style)
        btn_plus.clicked.connect(self._zoom_in)

        btn_reset = QPushButton("Ajuster")
        btn_reset.setFixedWidth(58)
        btn_reset.setStyleSheet(btn_style)
        btn_reset.clicked.connect(self._zoom_fit)

        self.zoom_lbl = QLabel("100%")
        self.zoom_lbl.setFont(QFont("Georgia", 9))
        self.zoom_lbl.setStyleSheet("color: #999;")
        self.zoom_lbl.setFixedWidth(42)

        hint = QLabel("Ctrl+molette pour zoomer  ·  cliquer-glisser pour déplacer")
        hint.setFont(QFont("Georgia", 8))
        hint.setStyleSheet("color: #bbb;")

        zoom_bar.addWidget(hint)
        zoom_bar.addStretch()
        zoom_bar.addWidget(btn_minus)
        zoom_bar.addWidget(self.zoom_lbl)
        zoom_bar.addWidget(btn_plus)
        zoom_bar.addWidget(btn_reset)
        layout.addLayout(zoom_bar)

        self.scroll = QScrollArea()
        self.scroll.setStyleSheet("border: none;")
        self.scroll.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.img_label = _PannableLabel(self.scroll)
        self.scroll.setWidget(self.img_label)
        layout.addWidget(self.scroll, stretch=1)

        # zoom initial : adapter à la largeur disponible (~500px)
        if not self._pixmap_orig.isNull():
            self._zoom = min(1.0, 500 / max(self._pixmap_orig.width(), 1))
        self._render()

        btn_close = QPushButton("Fermer")
        btn_close.setStyleSheet("""
            QPushButton {
                background: #1a1a1a; color: #fff;
                border: none; border-radius: 3px;
                padding: 10px; font-family: Georgia; font-size: 11px;
            }
            QPushButton:hover { background: #333; }
        """)
        btn_close.clicked.connect(self.close)
        layout.addWidget(btn_close)

    def _render(self):
        if self._pixmap_orig.isNull():
            self.img_label.setText("Image introuvable")
            return
        w = int(self._pixmap_orig.width() * self._zoom)
        h = int(self._pixmap_orig.height() * self._zoom)
        scaled = self._pixmap_orig.scaled(
            w, h,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        self.img_label.setPixmap(scaled)
        self.img_label.resize(scaled.width(), scaled.height())
        self.zoom_lbl.setText(f"{int(self._zoom * 100)}%")

    def _zoom_in(self):
        self._zoom = min(self._zoom + 0.25, 4.0)
        self._render()

    def _zoom_out(self):
        self._zoom = max(self._zoom - 0.25, 0.25)
        self._render()

    def _zoom_fit(self):
        if not self._pixmap_orig.isNull():
            self._zoom = min(1.0, (self.scroll.width() - 20) / max(self._pixmap_orig.width(), 1))
        self._render()

    def wheelEvent(self, event):
        # molette seule -> scroll normal (propagé au QScrollArea)
        # Ctrl+molette -> zoom
        if event.modifiers() & Qt.KeyboardModifier.ControlModifier:
            if event.angleDelta().y() > 0:
                self._zoom_in()
            else:
                self._zoom_out()
            event.accept()
        else:
            event.ignore()


class CorrectionDialog(QDialog):
    """Dialog with Nominatim-based autocomplete for addresses."""
    def __init__(self, initial: str = "", parent=None):
        super().__init__(parent)
        self.setWindowTitle("Corriger l'adresse — suggestions")
        self.setMinimumSize(500, 300)

        self.selected = None

        layout = QVBoxLayout(self)
        self.input = QLineEdit()
        self.input.setText(initial)
        layout.addWidget(self.input)

        btn_row = QHBoxLayout()
        self.btn_search = QPushButton("Rechercher")
        self.btn_search.clicked.connect(self._search)
        btn_row.addWidget(self.btn_search)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        from PyQt6.QtWidgets import QListWidget, QListWidgetItem
        self.listw = QListWidget()
        layout.addWidget(self.listw)

        foot = QHBoxLayout()
        self.btn_ok = QPushButton("Utiliser")
        self.btn_ok.clicked.connect(self._use_selected)
        self.btn_cancel = QPushButton("Annuler")
        self.btn_cancel.clicked.connect(self.reject)
        foot.addStretch()
        foot.addWidget(self.btn_ok)
        foot.addWidget(self.btn_cancel)
        layout.addLayout(foot)

        self.listw.itemDoubleClicked.connect(self._use_selected)

    def _search(self):
        q = self.input.text().strip()
        if not q:
            return
        url = "https://nominatim.openstreetmap.org/search"
        params = {"q": q, "format": "json", "addressdetails": 0, "limit": 6}
        try:
            resp = requests.get(url, params=params, timeout=10, headers={"User-Agent": "ReceiptReader/1.0"})
            resp.raise_for_status()
            data = resp.json()
        except Exception:
            data = []

        self.listw.clear()
        for item in data:
            display = item.get("display_name")
            if display:
                from PyQt6.QtWidgets import QListWidgetItem
                self.listw.addItem(QListWidgetItem(display))

    def _use_selected(self):
        it = self.listw.currentItem()
        if it:
            self.selected = it.text()
            self.accept()
        else:
            # fallback to input content
            txt = self.input.text().strip()
            if txt:
                self.selected = txt
                self.accept()


# ─────────────────────────────────────────────
# Zone de drop
# ─────────────────────────────────────────────
class DropZone(QFrame):
    files_dropped = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)
        self.setMinimumHeight(120)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._update_style(False)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(6)

        self.title = QLabel("Deposer les images ici")
        self.title.setFont(QFont("Georgia", 13))
        self.title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.title.setStyleSheet("color: #999; letter-spacing: 1px;")

        self.sub = QLabel("ou cliquer pour selectionner")
        self.sub.setFont(QFont("Georgia", 9))
        self.sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.sub.setStyleSheet("color: #bbb; letter-spacing: 0.5px;")

        layout.addWidget(self.title)
        layout.addWidget(self.sub)

    def _update_style(self, active: bool):
        border = "#333" if not active else "#111"
        bg = "#fafafa" if not active else "#f0f0f0"
        self.setStyleSheet(f"""
            DropZone {{
                border: 1.5px dashed {border};
                border-radius: 4px;
                background: {bg};
            }}
        """)

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()
            self._update_style(True)

    def dragLeaveEvent(self, e):
        self._update_style(False)

    def dropEvent(self, e):
        self._update_style(False)
        exts = (".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff", ".tif")
        paths = [u.toLocalFile() for u in e.mimeData().urls()
                 if u.toLocalFile().lower().endswith(exts)]
        if paths:
            self.files_dropped.emit(paths)

    def mousePressEvent(self, e):
        self.files_dropped.emit([])


# ─────────────────────────────────────────────
# Ligne résultat
# ─────────────────────────────────────────────
class ResultRow(QFrame):
    adresse_modifiee = pyqtSignal(str, str)

    CONFIANCE_STYLE = {
        "haute":   "#2d6a4f",  # vert   - fiable
        "moyenne": "#b9770e",  # orange - a verifier
        "basse":   "#c0392b",  # rouge  - a corriger
    }

    def _niveau(self, data):
        if data.get("statut") != "ok":
            return "basse"
        return data.get("confiance", "basse")

    def __init__(self, data: dict, adresse_labo: str):
        super().__init__()
        self.data = data
        self.adresse_labo = adresse_labo
        self.setStyleSheet("ResultRow { background: transparent; border-bottom: 1px solid #ececec; }")

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 10, 0, 10)
        layout.setSpacing(12)

        color = self.CONFIANCE_STYLE[self._niveau(data)]
        distance_failed = data.get("distance_raw") is None

        self.dot = QLabel("●")
        self.dot.setFixedWidth(14)
        self.dot.setFont(QFont("Arial", 10))
        self.dot.setStyleSheet(f"color: {color};")

        fichier = QLabel(data.get("fichier", ""))
        fichier.setFixedWidth(190)
        fichier.setFont(QFont("Georgia", 9))
        fichier.setStyleSheet("color: #999;")

        self.adresse_lbl = QLabel(data.get("adresse") or "—")
        self.adresse_lbl.setFont(QFont("Georgia", 10))
        self.adresse_lbl.setStyleSheet(f"color: {color};")
        self.adresse_lbl.setWordWrap(True)
        self.adresse_lbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

        self.dist_lbl = QLabel(data.get("distance_km") or "—")
        self.dist_lbl.setFixedWidth(70)
        self.dist_lbl.setFont(QFont("Georgia", 10))
        if data.get("distance_raw") is None:
            self.dist_lbl.setStyleSheet("color: #c0392b;")
        else:
            self.dist_lbl.setStyleSheet("color: #555;")
        self.dist_lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        btn_voir = QPushButton("Voir")
        btn_voir.setFixedWidth(44)
        btn_voir.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_voir.setStyleSheet("""
            QPushButton {
                background: transparent; color: #aaa;
                border: 1px solid #ddd; border-radius: 3px;
                padding: 4px 6px; font-family: Georgia; font-size: 10px;
            }
            QPushButton:hover { color: #1a1a1a; border-color: #999; }
        """)
        btn_voir.clicked.connect(self._voir_ticket)

        btn_edit = QPushButton("Corriger")
        btn_edit.setFixedWidth(60)
        btn_edit.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_edit.setStyleSheet("""
            QPushButton {
                background: transparent; color: #aaa;
                border: 1px solid #ddd; border-radius: 3px;
                padding: 4px 6px; font-family: Georgia; font-size: 10px;
            }
            QPushButton:hover { color: #1a1a1a; border-color: #999; }
        """)
        btn_edit.clicked.connect(self._corriger_adresse)

        layout.addWidget(self.dot)
        layout.addWidget(fichier)
        layout.addWidget(self.adresse_lbl)
        layout.addWidget(self.dist_lbl)
        layout.addWidget(btn_voir)
        layout.addWidget(btn_edit)

    def _voir_ticket(self):
        img_path = self.data.get("img_path", "")
        if img_path and os.path.exists(img_path):
            viewer = TicketViewer(img_path, self)
            viewer.exec()

    def _corriger_adresse(self):
        adresse_actuelle = self.data.get("adresse", "")
        dlg = CorrectionDialog(adresse_actuelle, self)
        if dlg.exec() and dlg.selected:
            nouvelle = dlg.selected.strip()
            self.data["adresse"] = nouvelle
            self.adresse_lbl.setText(nouvelle)

            dist, confiance = calculer_distance(self.adresse_labo, nouvelle)
            self.data["distance_raw"] = dist
            self.data["distance_km"] = f"{dist:.1f} km" if dist else "N/A"
            self.data["confiance"] = confiance
            self.data["statut"] = "ok"
            self.dist_lbl.setText(self.data["distance_km"])
            color = self.CONFIANCE_STYLE[self._niveau(self.data)]
            self.adresse_lbl.setStyleSheet(f"color: {color};")
            self.dot.setStyleSheet(f"color: {color};")
            self.adresse_modifiee.emit(self.data["fichier"], nouvelle)


# ─────────────────────────────────────────────
# Fenetre principale
# ─────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Maison Asteria Receipt Reader")
        self.setMinimumSize(1050, 700)
        self.image_paths = []
        self.all_results = []
        self.result_rows = []
        self.worker = None
        self.adresse_labo = ""
        self._build_ui()
        threading.Thread(target=get_reader, daemon=True).start()

    def _build_ui(self):
        self.setStyleSheet("""
            QMainWindow, QWidget { background: #ffffff; color: #1a1a1a; }
            QLineEdit {
                border: none; border-bottom: 1.5px solid #ddd;
                padding: 8px 0; font-family: Georgia; font-size: 13px;
                color: #1a1a1a; background: transparent;
            }
            QLineEdit:focus { border-bottom: 1.5px solid #1a1a1a; }
            QScrollArea { border: none; background: transparent; }
            QScrollBar:vertical { background: #fff; width: 4px; border: none; }
            QScrollBar::handle:vertical { background: #ddd; border-radius: 2px; }
            QProgressBar {
                border: none; background: #f0f0f0;
                height: 2px; border-radius: 1px;
            }
            QProgressBar::chunk { background: #1a1a1a; border-radius: 1px; }
        """)

        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Panneau gauche
        left = QWidget()
        left.setFixedWidth(280)
        left.setStyleSheet("background: #f7f7f7; border-right: 1px solid #e8e8e8;")
        ll = QVBoxLayout(left)
        ll.setContentsMargins(28, 36, 28, 36)
        ll.setSpacing(0)

        brand = QLabel("Maison Asteria\nReceipt Reader")
        brand.setFont(QFont("Georgia", 20))
        brand.setStyleSheet("color: #1a1a1a;")
        ll.addWidget(brand)

        ll.addSpacing(6)
        tagline = QLabel("Extraction automatique\nd'adresses de tickets")
        tagline.setFont(QFont("Georgia", 9))
        tagline.setStyleSheet("color: #999;")
        ll.addWidget(tagline)

        ll.addSpacing(40)

        lbl = QLabel("ADRESSE DU LABORATOIRE")
        lbl.setFont(QFont("Georgia", 8))
        lbl.setStyleSheet("color: #aaa; letter-spacing: 1.5px;")
        ll.addWidget(lbl)
        ll.addSpacing(8)

        self.labo_input = QLineEdit()
        self.labo_input.setPlaceholderText("ex: 15 rue de la Paix, 75001 Paris")
        ll.addWidget(self.labo_input)

        ll.addSpacing(40)

        self.btn_extract = QPushButton("Extraire les adresses")
        self.btn_extract.setEnabled(False)
        self.btn_extract.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_extract.setStyleSheet("""
            QPushButton {
                background: #1a1a1a; color: #fff; border: none;
                border-radius: 3px; padding: 13px 20px;
                font-family: Georgia; font-size: 12px; letter-spacing: 0.5px;
            }
            QPushButton:hover { background: #333; }
            QPushButton:disabled { background: #e0e0e0; color: #aaa; }
        """)
        self.btn_extract.clicked.connect(self._start)
        ll.addWidget(self.btn_extract)

        ll.addSpacing(10)

        # Export buttons: CSV and Excel (hidden until results ready)
        export_row = QWidget()
        export_layout = QHBoxLayout(export_row)
        export_layout.setContentsMargins(0, 0, 0, 0)
        export_layout.setSpacing(8)

        self.btn_export_csv = QPushButton("Exporter CSV")
        self.btn_export_csv.setVisible(False)
        self.btn_export_csv.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_export_csv.setStyleSheet("""
            QPushButton { background: #1a1a1a; color: #fff; border: none; border-radius: 3px; padding: 10px 14px; font-family: Georgia; font-size: 12px; }
            QPushButton:hover { background: #333; }
        """)
        self.btn_export_csv.clicked.connect(self._export_csv)

        self.btn_export_xlsx = QPushButton("Exporter Excel")
        self.btn_export_xlsx.setVisible(False)
        self.btn_export_xlsx.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_export_xlsx.setStyleSheet("""
            QPushButton { background: transparent; color: #1a1a1a; border: 1.5px solid #1a1a1a; border-radius: 3px; padding: 10px 14px; font-family: Georgia; font-size: 12px; }
            QPushButton:hover { background: #f0f0f0; }
        """)
        self.btn_export_xlsx.clicked.connect(self._export_xlsx)

        export_layout.addWidget(self.btn_export_csv)
        export_layout.addWidget(self.btn_export_xlsx)
        ll.addWidget(export_row)

        ll.addSpacing(10)

        self.btn_clear = QPushButton("Effacer")
        self.btn_clear.setVisible(False)
        self.btn_clear.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_clear.setStyleSheet("""
            QPushButton {
                background: transparent; color: #999; border: none;
                padding: 8px 0; font-family: Georgia; font-size: 11px; text-align: left;
            }
            QPushButton:hover { color: #1a1a1a; }
        """)
        self.btn_clear.clicked.connect(self._clear)
        ll.addWidget(self.btn_clear)

        ll.addStretch()

        self.status_label = QLabel("")
        self.status_label.setFont(QFont("Georgia", 9))
        self.status_label.setStyleSheet("color: #aaa;")
        self.status_label.setWordWrap(True)
        ll.addWidget(self.status_label)

        root.addWidget(left)

        # ── Panneau droit
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(40, 36, 40, 36)
        rl.setSpacing(0)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setTextVisible(False)
        self.progress.setFixedHeight(2)
        rl.addWidget(self.progress)
        rl.addSpacing(28)

        self.drop_zone = DropZone()
        self.drop_zone.files_dropped.connect(self._on_files_dropped)
        rl.addWidget(self.drop_zone)

        rl.addSpacing(28)

        self.section_label = QLabel("FICHIERS SELECTIONNES")
        self.section_label.setFont(QFont("Georgia", 8))
        self.section_label.setStyleSheet("color: #aaa; letter-spacing: 1.5px;")
        self.section_label.setVisible(False)
        rl.addWidget(self.section_label)

        rl.addSpacing(8)

        self.col_header = QWidget()
        ch = QHBoxLayout(self.col_header)
        ch.setContentsMargins(0, 0, 0, 0)
        ch.setSpacing(12)
        for txt, w in [("", 14), ("FICHIER", 190), ("ADRESSE EXTRAITE", None), ("DISTANCE", 70), ("", 44), ("", 60)]:
            lbl2 = QLabel(txt)
            lbl2.setFont(QFont("Georgia", 8))
            lbl2.setStyleSheet("color: #ccc; letter-spacing: 1px;")
            if w:
                lbl2.setFixedWidth(w)
            else:
                lbl2.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            ch.addWidget(lbl2)
        self.col_header.setVisible(False)
        rl.addWidget(self.col_header)

        self.legend = self._build_legend()
        self.legend.setVisible(False)
        rl.addWidget(self.legend)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setSpacing(0)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        scroll.setWidget(self.content_widget)
        rl.addWidget(scroll, stretch=1)

        root.addWidget(right, stretch=1)

    def _build_legend(self):
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(0, 6, 0, 6)
        lay.setSpacing(18)
        for color, text in [("#2d6a4f", "Fiable"), ("#b9770e", "À vérifier"), ("#c0392b", "À corriger")]:
            dot = QLabel("●")
            dot.setFont(QFont("Arial", 10))
            dot.setStyleSheet(f"color: {color};")
            lbl = QLabel(text)
            lbl.setFont(QFont("Georgia", 9))
            lbl.setStyleSheet("color: #777;")
            lay.addWidget(dot)
            lay.addWidget(lbl)
        lay.addStretch()
        return w

    def _on_files_dropped(self, paths):
        if not paths:
            paths, _ = QFileDialog.getOpenFileNames(
                self, "Selectionner des tickets", "",
                "Images (*.jpg *.jpeg *.png *.webp *.bmp *.tiff *.tif)"
            )
        if paths:
            self.image_paths = list(set(self.image_paths + paths))
            self._refresh_file_list()

    def _refresh_file_list(self):
        self._clear_content()
        self.section_label.setText(f"FICHIERS SELECTIONNES  —  {len(self.image_paths)} image(s)")
        self.section_label.setVisible(True)
        self.col_header.setVisible(False)
        self.legend.setVisible(False)
        self.btn_extract.setEnabled(True)
        self.btn_clear.setVisible(True)
        self.status_label.setText("")

    def _clear_content(self):
        self.result_rows = []
        while self.content_layout.count():
            item = self.content_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

    def _clear(self):
        self.image_paths = []
        self.all_results = []
        self._clear_content()
        self.section_label.setVisible(False)
        self.col_header.setVisible(False)
        self.legend.setVisible(False)
        self.btn_extract.setEnabled(False)
        self.btn_clear.setVisible(False)
        self.btn_export_csv.setVisible(False)
        self.btn_export_xlsx.setVisible(False)
        self.status_label.setText("")
        self.progress.setVisible(False)

    def _start(self):
        self.adresse_labo = self.labo_input.text().strip()
        if not self.adresse_labo:
            self.status_label.setText("Veuillez saisir l'adresse du laboratoire.")
            return

        if not ollama_server_reachable():
            QMessageBox.warning(
                self,
                "Ollama indisponible",
                "Ollama ne repond pas. Lance d'abord le serveur avec :\n\nollama serve"
            )
            return

        if not ollama_model_exists(MODEL):
            reply = QMessageBox.question(
                self,
                "Modele Ollama manquant",
                f"Le modele '{MODEL}' n'est pas installe.\n\nVoulez-vous le telecharger maintenant ?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )
            if reply != QMessageBox.StandardButton.Yes:
                self.status_label.setText(
                    f"Installe le modele avec : ollama pull {MODEL}"
                )
                return

            try:
                self.status_label.setText(f"Telechargement du modele {MODEL}...")
                QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
                ollama_pull_model(MODEL)
            except Exception as exc:
                QMessageBox.critical(self, "Echec du telechargement", str(exc))
                return
            finally:
                QApplication.restoreOverrideCursor()

        self._clear_content()
        self.all_results = []
        self.btn_extract.setEnabled(False)
        self.btn_export_csv.setVisible(False)
        self.btn_export_xlsx.setVisible(False)
        self.progress.setVisible(True)
        self.progress.setMaximum(len(self.image_paths))
        self.progress.setValue(0)

        self.section_label.setText("RESULTATS")
        self.section_label.setVisible(True)
        self.col_header.setVisible(True)
        self.legend.setVisible(True)

        self.worker = Worker(self.image_paths, self.adresse_labo)
        self.worker.progress.connect(self._on_progress)
        self.worker.result_ready.connect(self._on_result)
        self.worker.finished.connect(self._on_finished)
        self.worker.start()

    def _on_progress(self, i, filename):
        self.progress.setValue(i + 1)
        self.status_label.setText(f"Traitement : {filename}")

    def _on_result(self, data):
        self.all_results.append(data)
        row = ResultRow(data, self.adresse_labo)
        self.result_rows.append(row)
        self.content_layout.addWidget(row)

    def _on_finished(self):
        self.progress.setVisible(False)
        ok = sum(1 for r in self.all_results if r["statut"] == "ok")
        self.status_label.setText(f"{ok}/{len(self.all_results)} adresses trouvees")
        self.btn_extract.setEnabled(True)
        self.btn_export_csv.setVisible(True)
        self.btn_export_xlsx.setVisible(True)

    def _export_csv(self):
        if not self.all_results:
            return
        dossier = QFileDialog.getExistingDirectory(self, "Choisir le dossier de sauvegarde")
        if not dossier:
            return
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(dossier, f"rapport_csv_{timestamp}.csv")
            fieldnames = [
                "Adresse de départ",
                "Adresse d'arrivée",
                "Nom de l'entreprise",
                "Distance (km)",
                "A/R (km)",
                "Date",
                "Nom de l'image",
            ]
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
                writer.writeheader()
                for row in self.all_results:
                    dist = row.get("distance_raw")
                    writer.writerow({
                        "Adresse de départ": row.get("adresse", ""),
                        "Adresse d'arrivée": self.adresse_labo or "",
                        "Nom de l'entreprise": "",
                        "Distance (km)": f"{dist:.1f}" if dist else "",
                        "A/R (km)": f"{(dist*2):.1f}" if dist else "",
                        "Date": "",
                        "Nom de l'image": row.get("fichier", ""),
                    })
            self.status_label.setText(f"CSV exporté : {path}")
        except Exception as exc:
            QMessageBox.critical(self, "Erreur export CSV", str(exc))

    def _export_xlsx(self):
        if not self.all_results:
            return
        dossier = QFileDialog.getExistingDirectory(self, "Choisir le dossier de sauvegarde")
        if not dossier:
            return
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(dossier, f"rapport_excel_{timestamp}.xlsx")
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.worksheet.table import Table, TableStyleInfo
            except ImportError:
                import subprocess
                result = subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], capture_output=True)
                if result.returncode != 0:
                    raise RuntimeError("Impossible d'installer openpyxl. Installe-le manuellement avec : pip install openpyxl")
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.worksheet.table import Table, TableStyleInfo

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Notes de frais"

            title = "Notes de frais - Receipt Reader"
            ws.merge_cells("A1:G1")
            title_cell = ws["A1"]
            title_cell.value = title
            title_cell.font = Font(bold=True, name="Calibri", size=14, color="1F1F1F")
            title_cell.fill = PatternFill("solid", fgColor="EDEDED")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )

            headers = [
                "Adresse de départ",
                "Adresse d'arrivée",
                "Nom de l'entreprise",
                "Distance (km)",
                "A/R (km)",
                "Date",
                "Nom de l'image",
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F1F1F")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin", color="FFFFFF"),
                    right=Side(style="thin", color="FFFFFF"),
                    top=Side(style="thin", color="FFFFFF"),
                    bottom=Side(style="thin", color="FFFFFF"),
                )

            for row_data in self.all_results:
                dist = row_data.get("distance_raw")
                ws.append([
                    row_data.get("adresse", ""),
                    self.adresse_labo or "",
                    "",
                    f"{dist:.1f}" if dist else "",
                    f"{(dist*2):.1f}" if dist else "",
                    "",
                    row_data.get("fichier", ""),
                ])

            col_widths = [56, 56, 28, 14, 14, 12, 28]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            data_font = Font(name="Calibri", size=10)
            thin = Side(style="thin", color="D9D9D9")
            alt_fill = PatternFill("solid", fgColor="F7F7F7")
            white_fill = PatternFill("solid", fgColor="FFFFFF")
            fail_fill = PatternFill("solid", fgColor="FDE9E7")
            fail_font = Font(name="Calibri", size=10, color="9C0006")

            for row_index, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
                result = self.all_results[row_index - 3]
                adresse_failed = result.get("statut") != "ok"
                distance_failed = result.get("distance_raw") is None

                for cell in row:
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    if cell.column in (4, 5):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=thin,
                        right=thin,
                        top=thin,
                        bottom=thin,
                    )
                    cell.fill = alt_fill if row_index % 2 == 0 else white_fill

                if adresse_failed:
                    row[0].fill = fail_fill
                    row[0].font = fail_font

                if distance_failed:
                    row[3].fill = fail_fill
                    row[3].font = fail_font
                    row[4].fill = fail_fill
                    row[4].font = fail_font

            last_col = openpyxl.utils.get_column_letter(ws.max_column)
            last_row = ws.max_row
            table = Table(displayName="NotesDeFrais", ref=f"A2:{last_col}{last_row}")
            table_style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = table_style
            ws.add_table(table)

            ws.row_dimensions[1].height = 26
            ws.row_dimensions[2].height = 24
            for row_idx in range(3, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 22
            ws.freeze_panes = "A3"
            ws.sheet_view.showGridLines = False
            wb.save(path)
            self.status_label.setText(f"Excel exporté : {path}")
        except Exception as exc:
            QMessageBox.critical(self, "Erreur export Excel", str(exc))


def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Receipt Reader")
    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
